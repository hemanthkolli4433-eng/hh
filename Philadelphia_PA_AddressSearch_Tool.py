import os
import re
import sys
import csv
import time
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException


# =========================
# CONFIG
# =========================
START_URL = "https://property.phila.gov/"
OUTPUT_FILE_NAME = "Philadelphia_PA_AddressSearch_Tool_Output.xlsx"
INPUT_FILE_NAME = "PA.txt"

WAIT_TIME = 45
OWNER_PAGE_WAIT_TIME = 90
SAVE_EVERY = 1

HEADERS = [
    "Sno",
    "County",
    "State",
    "Input",
    "Property Address",
    "Property CSZ",
    "Owner name",
    "Parcel number",
    "Mailing address",
    "Mailing address CSZ",
]


# =========================
# PATH HELPERS
# =========================
def get_run_folder():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


RUN_FOLDER = get_run_folder()
OUTPUT_PATH = os.path.join(RUN_FOLDER, OUTPUT_FILE_NAME)
INPUT_PATH = os.path.join(RUN_FOLDER, INPUT_FILE_NAME)


# =========================
# TEXT HELPERS
# =========================
def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_multiline(value):
    if value is None:
        return ""

    lines = []
    seen = set()

    for line in str(value).replace("\r", "\n").split("\n"):
        line = clean_text(line)
        if not line:
            continue

        key = line.upper()
        if key not in seen:
            lines.append(line)
            seen.add(key)

    return "\n".join(lines)


def format_csz_pipe_state(value):
    """
    PHILADELPHIA, PA 19134-2301 -> PHILADELPHIA, |PA| 19134-2301
    Philadelphia PA 19124       -> Philadelphia |PA| 19124
    """
    value = clean_text(value)
    if not value:
        return ""

    match = re.match(
        r"^(?P<city>.*?)(?P<comma>,?)\s+(?P<state>[A-Za-z]{2})\s+(?P<zip>\d{5}(?:-\d{4})?)$",
        value,
    )

    if not match:
        return value

    city = clean_text(match.group("city"))
    comma = match.group("comma")
    state = match.group("state").upper()
    zip_code = match.group("zip")

    if comma:
        return f"{city}, |{state}| {zip_code}"

    return f"{city} |{state}| {zip_code}"


def extract_csz_from_text(text):
    """
    Finds:
        PHILADELPHIA, PA 19134-2301
        Philadelphia PA 19124
    """
    text = clean_text(text)
    if not text:
        return ""

    patterns = [
        r"PHILADELPHIA,\s*PA\s+\d{5}(?:-\d{4})?",
        r"PHILADELPHIA\s+PA\s+\d{5}(?:-\d{4})?",
        r"[A-Z][A-Z\s.'-]+,\s*PA\s+\d{5}(?:-\d{4})?",
        r"[A-Z][A-Z\s.'-]+\s+PA\s+\d{5}(?:-\d{4})?",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return clean_text(match.group(0)).upper()

    return ""


def read_input_values(file_path):
    values = []
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            for cell in row:
                text = clean_text(cell)
                if text:
                    values.append(text)
                    break

    elif ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                for cell in row:
                    text = clean_text(cell)
                    if text:
                        values.append(text)
                        break

    else:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            for line in f:
                text = clean_text(line)
                if text:
                    values.append(text)

    final_values = []
    seen = set()

    for value in values:
        key = value.upper()
        if key not in seen:
            final_values.append(value)
            seen.add(key)

    return final_values


# =========================
# EXCEL HELPERS
# =========================
def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Philadelphia Property"
    ws.append(HEADERS)
    return wb, ws


def auto_fit_columns(ws):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            for line in value.splitlines():
                max_len = max(max_len, len(line))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)


def save_workbook(wb, output_path):
    auto_fit_columns(wb.active)

    try:
        wb.save(output_path)
        return output_path

    except PermissionError:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        alt_path = os.path.join(
            RUN_FOLDER,
            f"Philadelphia_PA_Property_Output_{timestamp}.xlsx"
        )
        wb.save(alt_path)
        return alt_path


# =========================
# SELENIUM HELPERS
# =========================
def start_driver():
    options = Options()

    # GitHub Actions has no visible desktop, so Chrome must run headless there.
    # Local run stays visible unless you set HEADLESS=true.
    headless_mode = (
        os.environ.get("GITHUB_ACTIONS", "").lower() == "true"
        or os.environ.get("HEADLESS", "").lower() in ("1", "true", "yes")
    )

    if headless_mode:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")

    chrome_binary = (
        os.environ.get("CHROME_BIN")
        or os.environ.get("CHROME_PATH")
        or os.environ.get("GOOGLE_CHROME_BIN")
    )
    if chrome_binary and os.path.exists(chrome_binary):
        options.binary_location = chrome_binary

    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-software-rasterizer")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--remote-allow-origins=*")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(90)

    try:
        driver.maximize_window()
    except Exception:
        pass

    return driver


def wait_page_ready(driver, timeout=WAIT_TIME):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )


def find_visible_element(driver, selectors, timeout=WAIT_TIME):
    end_time = time.time() + timeout

    while time.time() < end_time:
        for selector in selectors:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)

            for element in elements:
                try:
                    if element.is_displayed():
                        return element
                except Exception:
                    pass

        time.sleep(0.5)

    raise TimeoutException(f"Element not found for selectors: {selectors}")


def find_search_input(driver):
    selectors = [
        "input#id-tagjks",
        "input[placeholder='Search the map']",
        "input.pvm-search-control-input",
        "input[class*='pvm-search-control-input']",
    ]
    return find_visible_element(driver, selectors, timeout=WAIT_TIME)


def set_vue_input_value(driver, input_element, value):
    driver.execute_script(
        """
        const input = arguments[0];
        const value = arguments[1];

        input.scrollIntoView({block: 'center'});
        input.focus();
        input.click();

        const nativeSetter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype,
            'value'
        ).set;

        nativeSetter.call(input, '');
        input.dispatchEvent(new Event('input', { bubbles: true }));
        input.dispatchEvent(new Event('change', { bubbles: true }));

        nativeSetter.call(input, value);
        input.dispatchEvent(new Event('input', { bubbles: true }));
        input.dispatchEvent(new Event('change', { bubbles: true }));
        """,
        input_element,
        value,
    )

    time.sleep(0.8)

    current_value = clean_text(input_element.get_attribute("value"))

    if clean_text(value).upper() not in current_value.upper():
        input_element.click()
        input_element.send_keys(Keys.CONTROL, "a")
        input_element.send_keys(Keys.BACKSPACE)
        time.sleep(0.2)

        for ch in value:
            input_element.send_keys(ch)
            time.sleep(0.02)

        time.sleep(0.8)


def click_search_button(driver):
    selectors = [
        "button[name='pvm-search-control-button']",
        "button[aria-label='search']",
        "button[title='search']",
        "button.pvm-search-control-button",
    ]

    button = find_visible_element(driver, selectors, timeout=WAIT_TIME)
    driver.execute_script("arguments[0].click();", button)


def close_modal_if_present(driver):
    try:
        close_selectors = [
            "button[aria-label='Close']",
            "button[aria-label='close']",
            "button[title='Close']",
            ".openmaps-modal-close",
            ".modal-close",
            ".close",
        ]

        for selector in close_selectors:
            buttons = driver.find_elements(By.CSS_SELECTOR, selector)

            for btn in buttons:
                try:
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        time.sleep(1)
                        return
                except Exception:
                    pass

        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        time.sleep(0.5)

    except Exception:
        pass


def owner_page_is_loaded(driver):
    """
    Main correction:
    Wait until pasted-code type page appears with Owner keyword/table.
    """
    try:
        return bool(
            driver.execute_script(
                """
                const modal = document.querySelector('.openmaps-modal-content');
                if (!modal) return false;

                const text = (modal.innerText || '').toLowerCase();

                const ownerTable =
                    modal.querySelector('table#ownerProperties.owner') ||
                    modal.querySelector('.owner-table table#ownerProperties');

                if (!ownerTable) return false;

                if (!text.includes('owner')) return false;
                if (!text.includes('opa account number')) return false;
                if (!text.includes('mailing address')) return false;

                const rows = ownerTable.querySelectorAll('tbody tr');
                return rows.length > 0;
                """
            )
        )
    except Exception:
        return False


def wait_for_owner_page(driver, timeout=OWNER_PAGE_WAIT_TIME):
    WebDriverWait(driver, timeout).until(lambda d: owner_page_is_loaded(d))


def capture_property_csz_from_visible_page(driver):
    """
    Try to capture Property CSZ from search result / visible page before modal scraping.
    This fixes missing:
        PHILADELPHIA, PA 19134-2301
    """
    try:
        text = driver.execute_script(
            """
            const candidates = [];

            const selectors = [
                '.pvm-search-results',
                '.pvm-search-result',
                '.pvm-search-result-item',
                '[role="listbox"]',
                '[role="option"]',
                '.search-results',
                '.search-result',
                '.mapboxgl-popup-content',
                '.openmaps-modal-header',
                '.openmaps-modal-content',
                'body'
            ];

            for (const selector of selectors) {
                document.querySelectorAll(selector).forEach(el => {
                    const rect = el.getBoundingClientRect();
                    const style = window.getComputedStyle(el);
                    const txt = (el.innerText || el.textContent || '').trim();

                    if (!txt) return;
                    if (style.visibility === 'hidden' || style.display === 'none') return;

                    candidates.push(txt);
                });
            }

            return candidates.join('\\n');
            """
        )

        return extract_csz_from_text(text)

    except Exception:
        return ""


def click_best_search_result_if_owner_not_loaded(driver, input_value):
    """
    If search button only shows result list, click the best matching result.
    Also returns result text so Property CSZ can be taken from it.
    """
    result = driver.execute_script(
        """
        const query = (arguments[0] || '').toLowerCase().trim();
        const queryParts = query.split(/\\s+/).filter(Boolean);
        const firstNumber = (query.match(/\\d+/) || [''])[0];

        const selectors = [
            '.pvm-search-results *',
            '.pvm-search-result',
            '.pvm-search-result-item',
            '[role="option"]',
            '.search-result',
            '.search-results *'
        ];

        let candidates = [];

        for (const selector of selectors) {
            document.querySelectorAll(selector).forEach(el => {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                const text = (el.innerText || el.textContent || '').trim();

                if (!text) return;
                if (rect.width <= 0 || rect.height <= 0) return;
                if (style.visibility === 'hidden' || style.display === 'none') return;

                let score = 0;
                const lower = text.toLowerCase();

                if (firstNumber && lower.includes(firstNumber)) score += 10;

                for (const part of queryParts) {
                    if (part.length >= 2 && lower.includes(part)) {
                        score += 1;
                    }
                }

                if (/philadelphia/i.test(text) && /pa/i.test(text)) score += 2;
                if (/\\d{5}(-\\d{4})?/.test(text)) score += 2;

                if (score > 0) {
                    candidates.push({ el, score, text });
                }
            });
        }

        candidates.sort((a, b) => b.score - a.score);

        if (candidates.length > 0) {
            candidates[0].el.scrollIntoView({block: 'center'});
            candidates[0].el.click();
            return candidates[0].text;
        }

        return '';
        """,
        input_value,
    )

    time.sleep(2)
    return clean_text(result)


# =========================
# SCRAPE HELPERS
# =========================
def extract_details_from_owner_page(driver, property_csz_hint=""):
    data = driver.execute_script(
        r"""
        function clean(value) {
            return (value || '')
                .replace(/\u00a0/g, ' ')
                .replace(/[ \t]+/g, ' ')
                .trim();
        }

        function uniqueLines(value) {
            const lines = (value || '')
                .replace(/\r/g, '\n')
                .split(/\n+/)
                .map(clean)
                .filter(Boolean);

            const out = [];
            const seen = new Set();

            for (const line of lines) {
                const key = line.toUpperCase();
                if (!seen.has(key)) {
                    out.push(line);
                    seen.add(key);
                }
            }

            return out;
        }

        function looksLikeCSZ(line) {
            line = clean(line);
            return /^[A-Z][A-Z\s.'-]+,?\s+[A-Z]{2}\s+\d{5}(?:-\d{4})?$/i.test(line);
        }

        function splitAddressAndCSZ(lines, forceLastLineAsCSZ) {
            const cleaned = [];
            const seen = new Set();

            for (const line of lines || []) {
                const value = clean(line);
                if (!value) continue;
                if (/^mailing address$/i.test(value)) continue;
                if (/^opa address$/i.test(value)) continue;

                const key = value.toUpperCase();
                if (!seen.has(key)) {
                    cleaned.push(value);
                    seen.add(key);
                }
            }

            if (cleaned.length === 0) {
                return { address: '', csz: '' };
            }

            if (cleaned.length === 1) {
                return { address: cleaned[0], csz: '' };
            }

            const lastLine = cleaned[cleaned.length - 1];

            // Important correction:
            // When address comes like 3 lines:
            //   123 MAIN ST
            //   APT 2
            //   PHILADELPHIA PA 19124
            // first 2 lines must stay in Address and last line must go to CSZ.
            if (forceLastLineAsCSZ || looksLikeCSZ(lastLine)) {
                return {
                    address: cleaned.slice(0, -1).join('\n'),
                    csz: lastLine
                };
            }

            return { address: cleaned.join('\n'), csz: '' };
        }

        const root = document.querySelector('.openmaps-modal-content') || document.body;

        function valueLinesByLabel(labelNames) {
            const wanted = labelNames.map(x => x.toLowerCase());

            const rows = Array.from(root.querySelectorAll('tr'));
            for (const row of rows) {
                const th = row.querySelector('th');
                const td = row.querySelector('td');

                if (!th || !td) continue;

                const label = clean(th.innerText || th.textContent).toLowerCase();

                for (const name of wanted) {
                    if (label === name || label.includes(name)) {
                        return uniqueLines(td.innerText || td.textContent);
                    }
                }
            }

            return [];
        }

        const opaAccountLines = valueLinesByLabel(['OPA Account Number']);
        const opaAddressLines = valueLinesByLabel(['OPA Address']);

        const opaAccount = opaAccountLines.length ? clean(opaAccountLines[0]) : '';

        // If OPA Address itself has 3 lines, split first lines as Property Address
        // and last line as Property CSZ. For 2 lines, split only if last line really looks like CSZ.
        const opaSplit = splitAddressAndCSZ(
            opaAddressLines,
            opaAddressLines.length >= 3
        );
        const opaAddress = opaSplit.address;
        const opaCSZ = opaSplit.csz;

        let ownerName = '';
        let parcelNumber = '';
        let mailingAddress = '';
        let mailingCSZ = '';

        const ownerTable =
            root.querySelector('table#ownerProperties.owner') ||
            root.querySelector('.owner-table table#ownerProperties');

        if (ownerTable) {
            const row = ownerTable.querySelector('tbody tr');

            if (row && row.cells.length >= 2) {
                const ownerCell = row.cells[0];
                const accountCell = row.cells[1];

                ownerName = uniqueLines(ownerCell.innerText || ownerCell.textContent).join('\n');

                const parcelSpan = accountCell.querySelector('span.large-owner');
                if (parcelSpan) {
                    parcelNumber = clean(parcelSpan.innerText || parcelSpan.textContent);
                }

                const lines = uniqueLines(accountCell.innerText || accountCell.textContent);

                let afterMailing = false;
                const mailingLines = [];

                for (const line of lines) {
                    if (!parcelNumber && /^\d{6,}$/.test(line)) {
                        parcelNumber = line;
                        continue;
                    }

                    if (/^mailing address$/i.test(line)) {
                        afterMailing = true;
                        continue;
                    }

                    if (afterMailing) {
                        mailingLines.push(line);
                    }
                }

                if (mailingLines.length > 0) {
                    // Mailing address block normally ends with CSZ.
                    // For 3-line mailing address, this gives first 2 lines as address
                    // and last line as CSZ.
                    const mailingSplit = splitAddressAndCSZ(
                        mailingLines,
                        mailingLines.length >= 2
                    );

                    mailingAddress = mailingSplit.address;
                    mailingCSZ = mailingSplit.csz;
                }
            }
        }

        const fullText = clean(root.innerText || root.textContent);

        return {
            opa_account: opaAccount,
            opa_address: opaAddress,
            opa_csz: opaCSZ,
            owner_name: ownerName,
            parcel_number: parcelNumber || opaAccount,
            mailing_address: mailingAddress,
            mailing_csz: mailingCSZ,
            full_text: fullText
        };
        """
    )

    if not isinstance(data, dict):
        data = {}

    property_csz = clean_text(property_csz_hint)

    if not property_csz:
        property_csz = clean_text(data.get("opa_csz", ""))

    if not property_csz:
        property_csz = extract_csz_from_text(data.get("full_text", ""))

    data["property_csz"] = property_csz
    return data


def make_output_row(serial_no, input_value, details):
    property_address = clean_multiline(details.get("opa_address", ""))
    if property_address:
        property_address = property_address.upper()

    property_csz = format_csz_pipe_state(details.get("property_csz", ""))
    mailing_address = clean_multiline(details.get("mailing_address", ""))
    mailing_csz = format_csz_pipe_state(details.get("mailing_csz", ""))

    return [
        serial_no,
        "Philadelphia",
        "PA",
        input_value,
        property_address,
        property_csz,
        clean_multiline(details.get("owner_name", "")),
        clean_text(details.get("parcel_number", "")),
        mailing_address,
        mailing_csz,
    ]


def make_blank_row(serial_no, input_value, remarks="No Record"):
    return [
        serial_no,
        "Philadelphia",
        "PA",
        input_value,
        "",
        "",
        "",
        remarks,
        "",
        "",
    ]


def scrape_one_input(driver, input_value):
    close_modal_if_present(driver)

    search_input = find_search_input(driver)
    set_vue_input_value(driver, search_input, input_value)

    click_search_button(driver)

    property_csz_hint = ""

    # Capture CSZ from search result/page text immediately after search click.
    time.sleep(2)
    property_csz_hint = capture_property_csz_from_visible_page(driver)

    # Main requirement: wait for pasted-code page with Owner keyword.
    try:
        wait_for_owner_page(driver, timeout=12)

    except TimeoutException:
        # Some searches show a result dropdown/list first.
        result_text = click_best_search_result_if_owner_not_loaded(driver, input_value)

        if not property_csz_hint:
            property_csz_hint = extract_csz_from_text(result_text)

        wait_for_owner_page(driver, timeout=OWNER_PAGE_WAIT_TIME)

    # After owner page appears, capture again in case header/body now has CSZ.
    if not property_csz_hint:
        property_csz_hint = capture_property_csz_from_visible_page(driver)

    details = extract_details_from_owner_page(driver, property_csz_hint=property_csz_hint)

    if not details.get("parcel_number"):
        raise Exception("Owner page appeared, but parcel number was not found")

    return details


# =========================
# MAIN
# =========================
def main():
    driver = None
    final_output_path = OUTPUT_PATH

    input_file = INPUT_PATH

    if not os.path.exists(input_file):
        print(f"ERROR: Input file not found: {input_file}")
        print("Create PA.txt in the same folder as this script/exe and run again.")
        return

    input_values = read_input_values(input_file)

    if not input_values:
        print(f"ERROR: Input file is empty: {input_file}")
        return

    print(f"Input file: {input_file}")
    print(f"Total input records: {len(input_values)}")
    print(f"Output file: {OUTPUT_PATH}")

    try:
        driver = start_driver()
        driver.get(START_URL)

        wait_page_ready(driver, timeout=90)
        find_search_input(driver)

        wb, ws = create_workbook()

        for index, input_value in enumerate(input_values, start=1):
            print(f"[{index}/{len(input_values)}] Searching: {input_value}", flush=True)

            try:
                details = scrape_one_input(driver, input_value)
                row = make_output_row(index, input_value, details)
                ws.append(row)

                print(f"Done: {input_value}", flush=True)

            except Exception as e:
                print(f"Error for {input_value}: {e}", flush=True)
                traceback.print_exc()

                # Helpful files for GitHub Actions debugging.
                try:
                    safe_index = str(index).zfill(4)
                    driver.save_screenshot(os.path.join(RUN_FOLDER, f"error_{safe_index}.png"))
                    with open(os.path.join(RUN_FOLDER, f"error_{safe_index}.html"), "w", encoding="utf-8") as f:
                        f.write(driver.page_source)
                except Exception:
                    pass

                ws.append(make_blank_row(index, input_value, "No Record"))

                try:
                    driver.get(START_URL)
                    wait_page_ready(driver, timeout=90)
                    find_search_input(driver)
                except Exception:
                    pass

            if index % SAVE_EVERY == 0:
                final_output_path = save_workbook(wb, final_output_path)

        final_output_path = save_workbook(wb, final_output_path)

        print("All records completed.")
        print(f"Excel saved here: {final_output_path}")

    except Exception as e:
        traceback.print_exc()
        print(f"ERROR: Tool stopped because of error: {e}")

        try:
            if driver:
                driver.save_screenshot(os.path.join(RUN_FOLDER, "fatal_error.png"))
                with open(os.path.join(RUN_FOLDER, "fatal_error.html"), "w", encoding="utf-8") as f:
                    f.write(driver.page_source)
        except Exception:
            pass

    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()